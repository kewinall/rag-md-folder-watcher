from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .base import BaseConverter
from ..utils import table_to_markdown


class XlsxConverter(BaseConverter):
    extensions = {".xlsx"}
    name = "xlsx"

    def convert(self, path: Path) -> str:
        workbook = load_workbook(str(path), read_only=True, data_only=True, keep_links=False)
        lines: list[str] = [f"# {path.stem}", ""]
        try:
            for sheet in workbook.worksheets:
                lines.append(f"## 工作表：{sheet.title}")
                rows = []
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if r_idx > self.config.excel_max_rows_per_sheet:
                        self.warnings.append(
                            f"工作表 {sheet.title} 已截斷，只保留前 {self.config.excel_max_rows_per_sheet} 列"
                        )
                        break
                    values = list(row[: self.config.excel_max_cols_per_sheet])
                    if any(v is not None and str(v).strip() for v in values):
                        rows.append(values)
                table_md = table_to_markdown(rows, self.config.excel_max_cols_per_sheet)
                if table_md:
                    lines.append(table_md)
                else:
                    lines.append("（空白工作表）\n")
        finally:
            workbook.close()
        return "\n".join(lines) + "\n"
